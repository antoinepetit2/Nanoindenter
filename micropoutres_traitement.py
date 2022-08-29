#!/usr/bin/env python
# coding: utf-8

# In[1]:


#import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import shutil

# import xlsxwriter module
import xlsxwriter
from openpyxl import load_workbook

########
######## define working directory
path='C:\\Users\\ap251712\\Desktop\\DT 2021-056_P5P6-nano\\'


#the file to read (for dimensions data)

os.chdir(path)
wb_to_read=load_workbook(filename = 'Resultats_Micropoutres.xlsx', data_only = True)
sheet_to_read = wb_to_read['Vierge_P5P6']
#column of the different index of the curves
first_column = sheet_to_read['C']


#reading a workbook with size of samples
workbook = xlsxwriter.Workbook('liste_valeurs.xlsx')
worksheet = workbook.add_worksheet()

row = 1
column = 0

worksheet.write('A1', 'data reference')
worksheet.write('B1', 'fmax measured')
worksheet.write('C1', 'displacement max')


#find reference file
filelist=os.listdir(path)
os.chdir(path)
for filename in filelist:
    if filename.startswith("ref"):
        ref_filename=filename
        
print('Found reference file as :')        
print(ref_filename)

#open reference file as reader mode
with open(ref_filename, 'r') as f:
    lines = f.readlines()
    lines = [line.rstrip() for line in lines]
    
#find where the interesting data start
for i in range(0,len(lines)):
    if lines[i].startswith('Time (s)'):
        break
        
# reading header of the file. Erase first lines        
print('start of data from reference curve :')
print(i)
del lines[0:i+2]

#read time, strength and depth (change into arrays)
time_ref = [float(line.split()[0]) for line in lines]
depth_ref = [float(line.split()[1]) for line in lines]
force_ref = [float(line.split()[2]) for line in lines]

#create a new folder for references curves    
ref_filename_folder=ref_filename.rstrip(".TXT") 

#je remove ça pour tester écrase et refait fichiers
#os.mkdir(ref_filename_folder)
#os.chdir(path+ref_filename_folder)

dir = path+ref_filename_folder
if os.path.exists(dir):
    shutil.rmtree(dir)
os.makedirs(dir)

os.chdir(path+ref_filename_folder)

#plot references curves 
    #strength vs time
plt.figure(1)
plt.plot(time_ref, force_ref, label='calibration test')
plt.ylabel('force (mN)')
plt.xlabel('time (s)')
plt.legend()
plt.grid()
plt.savefig('force_vs_time.png',bbox_inches='tight', transparent=True)
plt.close()

    #strength vs depth
plt.figure(2)
plt.plot(depth_ref, force_ref, label='calibration test')
plt.ylabel('force (mN)')
plt.xlabel('depth (nm)')
plt.legend()
plt.grid()
plt.savefig('force_vs_depth.png',bbox_inches='tight', transparent=True)
plt.close()

#identify the different parts of reference curve 
#penetration and thermal drift (td)

#calculate derivative of strength to detect different zones
from scipy.ndimage import gaussian_filter1d
len(time_ref)
time_minus=time_ref[:-1]
len(time_minus)
force_prime=np.diff(force_ref)
#smooth the derivative curve
smooth = gaussian_filter1d(force_prime, 100)

# to compensate the derivative function which remove one term in others arays
time_last=time_ref[-1]

    #derivative curves
plt.figure(3)
a=plt.plot(time_minus,force_prime, label='F derivativee')
b=plt.plot(time_minus,smooth,label='smoothed F derivative')
#c=plt.axhline(y=0.0020,label='Limit value for time selection',color='r')
#axs[1].plot(Tminus, smooth_d2 / np.max(smooth_d2), label='Second Derivative (scaled)')
plt.ylabel('force derivative (mN/s)')
plt.xlabel('time (s)')
plt.legend()
plt.savefig('force_derivative.png',bbox_inches='tight', transparent=True)
plt.close()
#plt.show()




# In[2]:


#Looking for the time until we stop increasing strength (i.e. derivative decreasing)

count=0
#strength stop increasing when derivative change sign
for i in smooth:
    if np.sign(i+1)!=np.sign(i):
    #if i <0.0020:
        print("Stop loop")
        break
    count +=1
    
print('max time for penetration correction is')
print(time_ref[count])

#create array that will be used for penetration correction
time_correct_pen=[]
force_correct_pen=[]
depth_correct_pen=[]

i=0
while i<count:
    time_correct_pen.append(time_ref[i])
    force_correct_pen.append(force_ref[i])
    depth_correct_pen.append(depth_ref[i])
    i+=1

    
from scipy.optimize import curve_fit
# Curve fitting function
def fit_func(x, a, b, c):
    return a*x**3+b*x**2+c*x

# Curve fitting during penetration testing
params = curve_fit(fit_func, depth_correct_pen, force_correct_pen)
[a, b, c] = params[0]
x_fit = np.linspace(depth_correct_pen[0], depth_correct_pen[-1], len(depth_correct_pen))
y_fit = a * x_fit**3 + b * x_fit**2 + c * x_fit

#plot original cruve and the fit
plt.figure(4)
plt.plot(depth_correct_pen, force_correct_pen, 'r', label='calibration test')         # Data
plt.plot(x_fit, y_fit, 'k', label='polynomial fit')  # Fitted curve
plt.text(x_fit[0],5.5,'y = %s *depth^3 +%s *depth^2 \n+%s *depth ' % (round(a,11), round(b,8), round(c,4)))
plt.legend()
plt.xlabel('depth (nm)')
plt.ylabel('force (mN)')
plt.grid()
plt.savefig('force_applied_fitting.png',bbox_inches='tight', transparent=True)
plt.close()
#plt.show()

#redefining with explicit names
y_fit_force_penetration=y_fit
x_fit_force_penetration=x_fit


# In[3]:


# Looking for thermal drift (TD)

# looking for time when we applied constant load (TD measurement)
time_array= np.array(time_ref)
bool_arr = time_array > time_ref[count+500]
#bool_arr = time_array > 220.0
time_min_td=np.where(bool_arr)[0]
time_min_td=time_min_td[0]
print('index time min for TD detection')
print(time_min_td)

index_time_td = np.where((smooth > -0.0005) & (smooth < 0.0005))
index_time_td = index_time_td[0] #break the tuple into array

# then exclude the first part of reference curve 
index_time_td = index_time_td[(index_time_td>=time_min_td)]
print('index of array where we observe TD')
print(index_time_td)

time_td=[]
depth_td=[]

w=len(index_time_td)
for i in range(0,w) :
    time_td.append(time_ref[index_time_td[i]])
    depth_td.append(depth_ref[index_time_td[i]])

if len(depth_td)==len(time_td):
    print ('Length of depth and time extracted for td are compatible')

from scipy import stats
slope, intercept, r_value, p_value, std_err = stats.linregress(time_td, depth_td)
# What we did is : depth_td(time_td)=slope*time_td+intercept


tmax=len(index_time_td)-1
t=np.arange(time_td[0],time_td[tmax],0.2)
plt.figure(5)
plt.plot(time_td,depth_td, label='penetration measured for constant strength')
plt.plot(t,(slope*t+intercept),label='linear interpolation')
plt.text(time_td[0],711.5,'y = %s *Time + %s\nr=%s' % (round(slope,4), round(intercept,5), round(r_value,4)))
plt.ylabel('penetration depth (nm)')
plt.xlabel('time (s)')
plt.legend()
plt.savefig('thermal_drift_fitting.png',bbox_inches='tight', transparent=True)
plt.close()
plt.show()


#end of study of reference curve
print (os.getcwd())

# In[4]:


#move to other datas
os.chdir(path)


print (os.getcwd())

#creating dictionary to store all data from text files
mytimes = dict()
myforces = dict()
mydepths = dict()
mydepths_td_corrected = dict()
mydepths_all_corrected = dict()
mystress=dict()


Nb_fichiers=0
for filename_data in filelist:
    if filename_data.endswith(".TXT") and filename_data.startswith("P5P6"): 
#il n'y a qu'un fichier comme ça
        Nb_fichiers+=1
        folder_data_name = filename_data.rstrip(".TXT")
        
        
        dir = path+folder_data_name
        if os.path.exists(dir):
            shutil.rmtree(dir)
        os.makedirs(dir)

        
        
        #☻os.mkdir(folder_data_name)
        #print(filename_data)
        appendix_number_data = folder_data_name.lstrip("P5P6")
        appendix_number_data = appendix_number_data.strip("#")
        #break
    
        appendix_number_data=int(appendix_number_data)

        print('new folder created is')
        print(folder_data_name)
        print('file being read is')
        print(filename_data)
        #print(Nb_fichiers)
        print('index of this file is')
        print(appendix_number_data)

        #creating dictionary to store all data from text files
        #mytimes = dict()
        #myforces = dict()
        #mydepths = dict()

        # reading the file
        with open(filename_data, 'r') as f:
            lines = f.readlines()
            lines = [line.rstrip() for line in lines]

        #select the starting line of datas
        for i in range(0,len(lines)):
            if lines[i].startswith('Time (s)'):
                break
        #print(i)
        debut_lignes=i+2
        del lines[0:debut_lignes]

        # extracting time, depth and strength from text file
        #time = [float(line.split()[0]) for line in lines]
        #depth = [float(line.split()[1]) for line in lines]
        #force = [float(line.split()[2]) for line in lines]
        mytimes[appendix_number_data] = [float(line.split()[0]) for line in lines]
        mydepths[appendix_number_data] = [float(line.split()[1]) for line in lines]
        myforces[appendix_number_data] = [float(line.split()[2]) for line in lines]

        #move to the new directory
        os.chdir(path+folder_data_name)
        cwd = os.getcwd()
        cwd

        #plot and save the datas

        # strength over time
        plt.figure()
        plt.plot(mytimes[appendix_number_data], myforces[appendix_number_data], label='force over time')
        plt.ylabel('force (mN)')
        plt.xlabel('time (s)')
        plt.legend()
        plt.grid()
        plt.savefig('force_vs_time.png',bbox_inches='tight', transparent=True)
        plt.close()

        #strength over depth
        plt.figure()
        plt.plot(mydepths[appendix_number_data], myforces[appendix_number_data], label='force over depth')
        plt.ylabel('force (mN)')
        plt.xlabel('depth (nm)')
        plt.legend()
        plt.grid()
        plt.savefig('force_vs_depth.png',bbox_inches='tight', transparent=True)
        plt.close()

        #strength derivative over time
        force_prime=np.diff(myforces[appendix_number_data])
        time_minus=mytimes[appendix_number_data]
        time_minus=time_minus[:-1]
        smooth = gaussian_filter1d(force_prime, 100)
        plt.figure()
        plt.plot(time_minus,force_prime, label='force derivative')
        #plt.plot(time_minus,smooth,label='Smoothed F_prime')
        plt.ylabel('force derivative')
        plt.xlabel('time (s)')
        plt.ylim(-0.2,0.2)
        plt.legend()
        plt.grid()
        plt.savefig('force_derivative.png',bbox_inches='tight', transparent=True)
        plt.close()

        #os.chdir(path) was used before mergin cells
        # found when the derivative change

        count3=0
        for i in force_prime:
            if i <-0.02 or i>0.02:
                print("On stoppe la boucle")
                break
            count3 +=1
        print('max index of interest in data is')
        print(count3)

        count_to_remove=len(mytimes[appendix_number_data])-count3


        #selection of the zone of interest in data

        del mytimes[appendix_number_data][-count_to_remove:]
        del myforces[appendix_number_data][-count_to_remove:]
        del mydepths[appendix_number_data][-count_to_remove:]

        # plot strength over displcament (raw data)

        plt.figure()
        plt.plot(mydepths[appendix_number_data],myforces[appendix_number_data], label='raw data')
        plt.ylabel('force (mN)')
        plt.xlabel('displacement uncorrected (nm)')
        plt.legend()
        plt.grid()
        plt.savefig('force_vs_displacement_uncorrected.png',bbox_inches='tight', transparent=True)
        plt.close()
        #plt.show()

        # first : correction of thermal drift

        depth_td_corrected=[]
        depth_interest_uncorrected=mydepths[appendix_number_data]
        time_interest=mytimes[appendix_number_data]
        i=0
        while i<len(mydepths[appendix_number_data]):
            depth_td_corrected.append(depth_interest_uncorrected[i]-slope*time_interest[i])
            i+=1

        mydepths_td_corrected[appendix_number_data]=depth_td_corrected

  

      # plot of thermal drift (comparison)
        ind_max=len(mydepths[appendix_number_data])-1
        ind_min=int(0.8*ind_max)
        xmax=time_interest[ind_max]+0.1*time_interest[ind_max]
        xmin=time_interest[ind_min]
        
        ymin=depth_td_corrected[ind_min]
        ymax=depth_td_corrected[ind_max]+0.1*depth_td_corrected[ind_max]
       

 
        
        plt.figure(figsize=(10, 3))
        plt.subplot(121)
        plt.plot(mytimes[appendix_number_data],mydepths[appendix_number_data],label='displacement without TD correction')
        plt.plot(mytimes[appendix_number_data],mydepths_td_corrected[appendix_number_data],label='displacement with TD correction')
        plt.ylabel('displacement (nm)')
        plt.xlabel('time (s)')
        plt.legend()
        plt.grid()
        plt.subplot(122)
        plt.plot(mytimes[appendix_number_data],mydepths[appendix_number_data],label='displacement without TD correction')
        plt.plot(mytimes[appendix_number_data],mydepths_td_corrected[appendix_number_data],label='displacement with TD correction')
        plt.ylabel('displacement (nm)')
        plt.xlabel('time (s)')
        plt.axis([xmin, xmax, ymin, ymax])
        plt.legend()
        plt.grid()
        plt.savefig('TD_correction_plot.png',bbox_inches='tight', transparent=True)
        plt.close()
        #plt.show()

        #plot of comparison of strength with and without TD consideration

        plt.plot(mydepths[appendix_number_data],myforces[appendix_number_data],label='force without TD correction')
        plt.plot(mydepths_td_corrected[appendix_number_data],myforces[appendix_number_data],label='force with TD correction')
        plt.ylabel('force (mN)')
        plt.xlabel('displacement (nm)')
        plt.legend()
        plt.grid()
        plt.savefig('force_with_td_correction.png',bbox_inches='tight', transparent=True)
        plt.close()
        #plt.show()

        #plot of strength with td correction and curve fitting of penetration test

        plt.plot(x_fit_force_penetration,y_fit_force_penetration, label='force during penetration test')
        plt.plot(mydepths_td_corrected[appendix_number_data],myforces[appendix_number_data],label='force with TD correction')
        #plt.plot(Hcorrect,smooth, label='Smoothed Fprime')
        plt.legend()
        plt.grid()
        plt.ylabel('force (mN)')
        plt.xlabel('depth (nm)')
        plt.savefig('force_with_td_correction_vs_penetrations_ref.png',bbox_inches='tight', transparent=True)
        plt.close()
        #plt.show()

        # correction of penetration (by uing fitting equation determined before)

        depth_corrected_all=[]
        force_interest_uncorrected=myforces[appendix_number_data]
        i=0
        while i<len(mydepths_td_corrected[appendix_number_data]):
            tau=min(range(len(y_fit_force_penetration)), key=lambda v: abs(y_fit_force_penetration[v]-force_interest_uncorrected[i]))
            depth_corrected_all.append(depth_td_corrected[i]-x_fit_force_penetration[tau])
            i+=1
        #print(len(depth_corrected_all))

        mydepths_all_corrected[appendix_number_data]=depth_corrected_all

        # plot of curves with and without penetration consideration

        #plt.plot(depth_correct_pen,force_correct_pen, label='F_indentation')
        plt.plot(x_fit_force_penetration,y_fit_force_penetration, label='calibration')
        plt.plot(mydepths_td_corrected[appendix_number_data],myforces[appendix_number_data], label='test (no correction)')
        plt.plot(mydepths_all_corrected[appendix_number_data],myforces[appendix_number_data], label='test (with correction)')
        plt.legend()
        plt.grid()
        plt.ylabel('force (mN)')
        plt.xlabel('displacement (nm)')
        #plt.axis([0, 1850, 0, 3])
        plt.savefig('strength_vs_displacement_.png',bbox_inches='tight', transparent=True)
        plt.close()
        #plt.show()

        #I read max force reached and max depth reached
        
        worksheet.write(row, column, appendix_number_data)
        
        max_force = np.max(myforces[appendix_number_data])
        print('maximum force measured')
        print(max_force)
        worksheet.write(row, column+1, max_force)
        
        
        max_defo=np.max(mydepths_all_corrected[appendix_number_data])
        print('maximum displacement measured')
        print(max_defo)
        worksheet.write(row, column+2, max_defo)
        
        #need to be able to read all txt files
        row+=1
        
        os.chdir(path)
    
        stress_calculated=[]
        LenP5P6=len(first_column)
        print(LenP5P6)
        i=0
        while i<LenP5P6:
            if (first_column[i].value)== appendix_number_data:
                print('trouve une correspondance')
                print (i)
        #find the quadratic moment and center of gravity of excel sheet
                L_calculated = sheet_to_read.cell(row=i+1, column=18).internal_value
                z_calculated = sheet_to_read.cell(row=i+1, column=19).internal_value
                I_calculated = sheet_to_read.cell(row=i+1, column=20).internal_value
                print ('valeur de I')
                print(I_calculated)
                k=0
                while k<len(myforces[appendix_number_data]):
                    stress_calculated.append((force_interest_uncorrected[k]*0.001*L_calculated*z_calculated)/(I_calculated))
                    k+=1
                mystress[appendix_number_data]=stress_calculated
                break
            else:
                i+=1
        os.chdir(path+folder_data_name)  
        
#        plt.plot(x_fit_force_penetration,y_fit_force_penetration, label='stress during penetration test')
#        plt.plot(mydepths_td_corrected[appendix_number_data],myforces[appendix_number_data], label='analytical stress')
        plt.plot(mydepths_all_corrected[appendix_number_data],mystress[appendix_number_data], label='analytical stress')
        plt.legend()
        plt.grid()
        plt.ylabel('stress (N/m²)')
        plt.xlabel('displacement (nm)')
        #plt.axis([0, 1850, 0, 3])
        plt.savefig('stress_vs_displacement_.png',bbox_inches='tight', transparent=True)
        plt.close()
        #plt.show()
        
        os.chdir(path)
print('Number of file created is')   
print(Nb_fichiers)

workbook.close()

######################################################################
########## Try to save dictionaries ####################
#reuse of data later (simulations comparison)


import json

dir_sauv_dico="Experimental"
dir = path+dir_sauv_dico
if os.path.exists(dir):
    shutil.rmtree(dir)
os.makedirs(dir)

os.chdir(path+dir_sauv_dico)


times_file=open("times_file", "w")
json.dump(mytimes, times_file)
times_file.close()

forces_file=open("forces_file", "w")
json.dump(myforces, forces_file)
forces_file.close()

depths_file=open("depths_file", "w")
json.dump(mydepths_all_corrected, depths_file)
depths_file.close()

stress_file=open("stress_file", "w")
json.dump(mystress, stress_file)
stress_file.close()


os.chdir(path)
###################  define a custom plot function  ##################
# use it to plot several samples on one figure

def custom_plot_force(list_arg, list_label):
    plt.figure()
    Nb=len(list_arg)
    i=0
    while i<Nb:
        plt.plot(mydepths_all_corrected[list_arg[i]],myforces[list_arg[i]], label=list_label[i])
        i+=1
    plt.legend()
    plt.grid()
    plt.ylabel('force (mN)')
    plt.xlabel('displacement corrected(nm)')
    plt.show()

def custom_plot_stress(list_arg, list_label):
    plt.figure()
    Nb=len(list_arg)
    i=0
    while i<Nb:
        plt.plot(mydepths_all_corrected[list_arg[i]],mystress[list_arg[i]], label=list_label[i])
        i+=1
    plt.legend()
    plt.grid()
    plt.ylabel('stress (N/m²)')
    plt.xlabel('displacement corrected (nm)')
    plt.show()



